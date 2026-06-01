import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import matplotlib.pyplot as plt
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 1. CALCUL DES AGRÉGATS
# ─────────────────────────────────────────────────────────────────────────────
 
# ─────────────────────────────────────────────────────────────────────────────
# DICTIONNAIRE DE CORRESPONDANCE cat_assure → type_beneficiaire
# ─────────────────────────────────────────────────────────────────────────────

CAT_ASSURE_TO_TYPE_BENEFICIAIRE = {
    # Assurés principaux
    "Assurés":      "Assuré",
    "assure":           "Assuré",
    "ASS":          "Assuré",
    "AS":          "Assuré",
    "ASSURÉ":       "Assuré",
    "ASSURÉS":      "Assuré",
    "ASSURES":      "Assuré",
    "Principal":    "Assuré",
    "Actif":        "Assuré",
    "Actifs":       "Assuré",
    "Salarié":      "Assuré",
    "Salariés":     "Assuré",
    "Retraité":     "Assuré",
    "Retraités":    "Assuré",
    "Portabilité":  "Assuré",
    "Port":         "Assuré",

    # Conjoints
    "conjoint":     "Conjoint",
    "CJ":           "Conjoint",
    "Cjs":         "Conjoint",
    "CONJOINT":     "Conjoint",
    "CONJOINTE":    "Conjoint",
    "Conjoints":    "Conjoint",
    "ConjointS":     "Conjoint",
    "Conjointe":    "Conjoint",
    "Conjoints":    "Conjoint",
    "Époux":        "Conjoint",
    "Épouse":       "Conjoint",
    "Concubin":     "Conjoint",
    "Concubine":    "Conjoint",
    "PACS":         "Conjoint",

    # Enfants
    "enfant":       "Enfant",
    "EN":           "Enfant",
    "autre":        "Enfant",
    "ENFANT":       "Enfant",
    "ENFANTS":      "Enfant",
    "Enfant":       "Enfant",
    "Enfants":      "Enfant",
    "EF":           "Enfant",
    "Autres":       "Enfant",
    "Ayants droit": "Enfant",
    "AD":           "Enfant",
}


def normaliser_type_beneficiaire(df, col_source="type_beneficiaire",
                                  col_cible="type_beneficiaire"):
                            
    
    df = df.copy()
    serie_source = df[col_source].astype(str).str.strip()
    mapped = serie_source.map(CAT_ASSURE_TO_TYPE_BENEFICIAIRE)
    # Là où le mapping n'a pas trouvé de correspondance, garder la valeur originale
    df[col_cible] = mapped.where(mapped.notna(), serie_source)
    return df



def depenses_santé_famille_acte_type_beneficiaire(df):

    
    ordre_famille = [
        "Hospitalisation",
        "Consultations et visites",
        "Soins courants",
        "Pharmacie",
        "Dentaire",
        "Optique",
    ]
 
    df = df.copy()
    df = df[df['famille_acte_aops'].isin(ordre_famille)]
    df['famille_acte_aops'] = pd.Categorical(
        df['famille_acte_aops'], categories=ordre_famille, ordered=True
    )
    
 
    annees = sorted(df['annee_survenance'].unique())
    print(f"Années de survenance trouvées : {annees}")
    assert len(annees) == 2, "La fonction attend exactement 2 années."
    an1, an2 = annees


    
 
    def evol(a, b):
        return (b - a) / a if a else np.nan
 
    beneficiaires = sorted(df['type_beneficiaire'].unique())
    blocks = []
 
    for ben in beneficiaires:
        df_ben = df[df['type_beneficiaire'] == ben]
 
        rc = df_ben.pivot_table(values='RC', index='famille_acte_aops',
                                columns='annee_survenance', aggfunc='sum')
        nb = df_ben.pivot_table(values='id_beneficiaire', index='famille_acte_aops',
                                columns='annee_survenance', aggfunc='nunique')
 
        for col in [an1, an2]:
            if col not in rc.columns: rc[col] = np.nan
            if col not in nb.columns: nb[col] = np.nan
 
        rc  = rc[[an1, an2]].reindex(ordre_famille)
        nb  = nb[[an1, an2]].reindex(ordre_famille)
        moy = rc / nb
 
        detail = pd.DataFrame({
            ('Consommants', an1):               nb[an1],
            ('Consommants', an2):               nb[an2],
            ('Consommants', f'Evol {an2}/{an1}'): (nb[an2] - nb[an1]) / nb[an1].replace(0, np.nan),
            ('RC',          an1):               rc[an1],
            ('RC',          an2):               rc[an2],
            ('RC',          f'Evol {an2}/{an1}'): (rc[an2] - rc[an1]) / rc[an1].replace(0, np.nan),
            ('RC_moyen',    an1):               moy[an1],
            ('RC_moyen',    an2):               moy[an2],
            ('RC_moyen',    f'Evol {an2}/{an1}'): (moy[an2] - moy[an1]) / moy[an1].replace(0, np.nan),
        })
        detail.columns = pd.MultiIndex.from_tuples(detail.columns)
        detail.index = pd.MultiIndex.from_tuples(
            [(ben, f) for f in detail.index],
            names=['type_beneficiaire', 'famille_acte_aops']
        )
 
        rc_tot  = {an1: rc[an1].sum(),  an2: rc[an2].sum()}
        nb_tot  = {
            an1: df_ben[df_ben['annee_survenance'] == an1]['id_beneficiaire'].nunique(),
            an2: df_ben[df_ben['annee_survenance'] == an2]['id_beneficiaire'].nunique(),
        }
        moy_tot = {yr: rc_tot[yr] / nb_tot[yr] if nb_tot[yr] else np.nan for yr in [an1, an2]}
 
        tot_row = pd.DataFrame([{
            ('Consommants', an1):               nb_tot[an1],
            ('Consommants', an2):               nb_tot[an2],
            ('Consommants', f'Evol {an2}/{an1}'): evol(nb_tot[an1],  nb_tot[an2]),
            ('RC',          an1):               rc_tot[an1],
            ('RC',          an2):               rc_tot[an2],
            ('RC',          f'Evol {an2}/{an1}'): evol(rc_tot[an1],  rc_tot[an2]),
            ('RC_moyen',    an1):               moy_tot[an1],
            ('RC_moyen',    an2):               moy_tot[an2],
            ('RC_moyen',    f'Evol {an2}/{an1}'): evol(moy_tot[an1], moy_tot[an2]),
        }])
        tot_row.columns = pd.MultiIndex.from_tuples(tot_row.columns)
        tot_row.index = pd.MultiIndex.from_tuples(
            [(ben, 'Total')], names=['type_beneficiaire', 'famille_acte_aops']
        )
        blocks.append(pd.concat([detail, tot_row]))
 
    rc_g  = {yr: df[df['annee_survenance'] == yr]['RC'].sum() for yr in annees}
    nb_g  = {yr: df[df['annee_survenance'] == yr]['id_beneficiaire'].nunique() for yr in annees}
    moy_g = {yr: rc_g[yr] / nb_g[yr] if nb_g[yr] else np.nan for yr in annees}
 
    global_row = pd.DataFrame([{
        ('Consommants', an1):               nb_g[an1],
        ('Consommants', an2):               nb_g[an2],
        ('Consommants', f'Evol {an2}/{an1}'): evol(nb_g[an1],  nb_g[an2]),
        ('RC',          an1):               rc_g[an1],
        ('RC',          an2):               rc_g[an2],
        ('RC',          f'Evol {an2}/{an1}'): evol(rc_g[an1],  rc_g[an2]),
        ('RC_moyen',    an1):               moy_g[an1],
        ('RC_moyen',    an2):               moy_g[an2],
        ('RC_moyen',    f'Evol {an2}/{an1}'): evol(moy_g[an1], moy_g[an2]),
    }])
    global_row.columns = pd.MultiIndex.from_tuples(global_row.columns)
    global_row.index = pd.MultiIndex.from_tuples(
        [('Total', '')], names=['type_beneficiaire', 'famille_acte_aops']
    )
 
    return pd.concat(blocks + [global_row])
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 2. EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────
 
# Palette couleurs par bénéficiaire  →  (cellule gauche, fond lignes, fond total)
PALETTE = {
    "Assuré":   ("173A64", "DCE6F1", "BDD7EE"),
    "Conjoint": ("662064", "EAD9F5", "D6B4F0"),
    "Enfant":   ("D86173", "FAE0E8", "F4B8CE"),
    # Fallbacks si les valeurs diffèrent légèrement
    "Assurés":   ("173A64", "DCE6F1", "BDD7EE"),
    "Conjoints": ("662064", "EAD9F5", "D6B4F0"),
    "Enfant":   ("D86173", "FAE0E8", "F4B8CE"),
}
 
C_HEADER_DARK = "173A64"
C_HEADER_MID  = "2E5090"
C_EVOL_HDR    = "EE9744"
C_EVOL_BG     = "FCEAD8"
C_TOTAL_ROW   = "2A0C53"
C_WHITE       = "FFFFFF"
C_BLACK       = "000000"
C_RED_NEG     = "173A64"
C_GREEN_POS   = "173A64"
 
 
def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)
 
def _font(bold=False, color=C_BLACK, size=9):
    return Font(bold=bold, color=color, size=size, name="HelveticaNeue")
 
def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)
 
def _left():
    return Alignment(horizontal="left", vertical="center")
 
_thin  = Side(style="thin",   color="AAAAAA")
_thick = Side(style="medium", color="555555")
 
def _border():        return Border(left=_thin,  right=_thin,  top=_thin, bottom=_thin)
def _border_bold():   return Border(left=_thin,  right=_thin,  top=_thin, bottom=_thick)
 
 
def _fmt_pct(v):
    """Formate un ratio flottant en chaîne '42%' ou '-5%'."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "N/A"
    return f"{round(v * 100)}%"
 
def _fmt_num(v, decimals=0):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    return round(v, decimals)
 
NUM_FMT = '#,##0'   # separateur de milliers, 0 decimale

def export_excel(table: pd.DataFrame, out_path: str):
    """
    Prend le DataFrame retourné par depenses_santé_famille_acte_type_beneficiaire()
    et génère un fichier Excel formaté à out_path.
    """
    # Déduire les deux années depuis les colonnes MultiIndex
    annees = sorted({c[1] for c in table.columns if isinstance(c[1], (int, np.integer))})
    an1, an2 = annees
    evol_label = f'Evol {an2}/{an1}'
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Étude par survenance"
 
    # ── Largeurs colonnes ────────────────────────────────────────────────────
    for col_letter, width in zip("ABCDEFGHIJK", [12, 22, 8, 8, 9, 11, 11, 9, 8, 8, 9]):
        ws.column_dimensions[col_letter].width = width
 
    # ── Ligne 1 : titres groupes de colonnes ────────────────────────────────
    ws.merge_cells("A1:B2")
    ws["A1"] = "Étude par survenance\nDonnées arrêtées au 31/12/N"
    ws["A1"].font = _font(bold=True, color=C_WHITE)
    ws["A1"].fill = _fill(C_HEADER_DARK)
    ws["A1"].alignment = _center()
 
    for cells, label, color in [
        ("C1:D1", "Consommants",                        C_HEADER_DARK),
        ("F1:G1", "Remboursement complémentaire",       C_HEADER_DARK),
        ("I1:J1", "Consommation moyenne par consommant",C_HEADER_DARK),
    ]:
        ws.merge_cells(cells)
        c = ws[cells.split(":")[0]]
        c.value = label
        c.font  = _font(bold=True, color=C_WHITE)
        c.fill  = _fill(color)
        c.alignment = _center()
 
    for col in ["E1", "H1", "K1"]:
        ws[col] = f"Évolution\n{an2}/{an1}"
        ws[col].font      = _font(bold=True, color=C_WHITE)
        ws[col].fill      = _fill(C_EVOL_HDR)
        ws[col].alignment = _center()
 
    # ── Ligne 2 : sous-headers années ───────────────────────────────────────
    for col, label in [("C", an1), ("D", an2), ("F", an1), ("G", an2), ("I", an1), ("J", an2)]:
        ws[f"{col}2"] = label
        ws[f"{col}2"].font      = _font(bold=True, color=C_WHITE)
        ws[f"{col}2"].fill      = _fill(C_HEADER_MID)
        ws[f"{col}2"].alignment = _center()
 
    for col in ["E2", "H2", "K2"]:
        ws[col].fill = _fill(C_EVOL_HDR)
 
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16
 
    # ── Écriture des lignes ──────────────────────────────────────────────────
    row_idx   = 3
    ben_list  = [b for b in table.index.get_level_values(0).unique() if b != "Total"]
 
    for ben in ben_list:
        c_left, c_row, c_tot = PALETTE.get(ben, ("444444", "EEEEEE", "CCCCCC"))
        df_ben = table.loc[ben]
        familles = [f for f in df_ben.index if f != "Total"]
        start_row = row_idx
 
        for famille in familles:
            r = df_ben.loc[famille]
 
            vals = [
                None,
                famille,
                _fmt_num(r[('Consommants', an1)]),
                _fmt_num(r[('Consommants', an2)]),
                _fmt_pct(r[('Consommants', evol_label)]),
                _fmt_num(r[('RC', an1)]),
                _fmt_num(r[('RC', an2)]),
                _fmt_pct(r[('RC', evol_label)]),
                _fmt_num(r[('RC_moyen', an1)]),
                _fmt_num(r[('RC_moyen', an2)]),
                _fmt_pct(r[('RC_moyen', evol_label)]),
            ]
 
            for col_i, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=col_i, value=val)
                cell.fill      = _fill(c_row)
                cell.font      = _font(size=9)
                cell.border    = _border()
                cell.alignment = _left() if col_i == 2 else _center()
                if col_i in [3, 4, 6, 7, 9, 10]:
                    cell.number_format = NUM_FMT
                if col_i in [5, 8, 11]:
                    cell.fill = _fill(C_EVOL_BG)
                    neg = isinstance(val, str) and val.startswith("-")
                    cell.font = _font(size=9, color=C_RED_NEG if neg else C_GREEN_POS)
            row_idx += 1
 
        # Cellule groupe fusionnée (colonne A) — couvre les lignes familles seulement
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=row_idx - 1, end_column=1)
        g_cell = ws.cell(row=start_row, column=1)
        g_cell.value     = ben
        g_cell.fill      = _fill(c_left)
        g_cell.font      = _font(bold=True, color=C_WHITE)
        g_cell.alignment = _center()
 
        # Ligne total bénéficiaire
        rt = df_ben.loc["Total"]
        tot_vals = [
            None, None,
            _fmt_num(rt[('Consommants', an1)]),
            _fmt_num(rt[('Consommants', an2)]),
            _fmt_pct(rt[('Consommants', evol_label)]),
            _fmt_num(rt[('RC', an1)]),
            _fmt_num(rt[('RC', an2)]),
            _fmt_pct(rt[('RC', evol_label)]),
            _fmt_num(rt[('RC_moyen', an1)]),
            _fmt_num(rt[('RC_moyen', an2)]),
            _fmt_pct(rt[('RC_moyen', evol_label)]),
        ]
        for col_i, val in enumerate(tot_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_i, value=val)
            cell.fill      = _fill(c_tot)
            cell.font      = _font(bold=True, size=9)
            cell.border    = _border_bold()
            cell.alignment = _center()
            if col_i in [3, 4, 6, 7, 9, 10]:
                cell.number_format = NUM_FMT
            if col_i in [5, 8, 11]:
                neg = isinstance(val, str) and val.startswith("-")
                cell.font = _font(bold=True, size=9, color=C_RED_NEG if neg else C_GREEN_POS)
        row_idx += 1
 
    # ── Ligne Total global ───────────────────────────────────────────────────
    gt = table.loc["Total"].iloc[0]
 
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    ws.cell(row=row_idx, column=1).value     = "Total"
    ws.cell(row=row_idx, column=1).fill      = _fill(C_TOTAL_ROW)
    ws.cell(row=row_idx, column=1).font      = _font(bold=True, color=C_WHITE)
    ws.cell(row=row_idx, column=1).alignment = _center()
    ws.cell(row=row_idx, column=1).number_format = NUM_FMT
 
    global_vals = [
        None, None,
        _fmt_num(gt[('Consommants', an1)]),
        _fmt_num(gt[('Consommants', an2)]),
        _fmt_pct(gt[('Consommants', evol_label)]),
        _fmt_num(gt[('RC', an1)]),
        _fmt_num(gt[('RC', an2)]),
        _fmt_pct(gt[('RC', evol_label)]),
        _fmt_num(gt[('RC_moyen', an1)]),
        _fmt_num(gt[('RC_moyen', an2)]),
        _fmt_pct(gt[('RC_moyen', evol_label)]),
    ]
    for col_i, val in enumerate(global_vals, start=1):
        if col_i <= 2:
            continue
        cell = ws.cell(row=row_idx, column=col_i, value=val)
        cell.fill      = _fill(C_TOTAL_ROW)
        cell.font      = _font(bold=True, color=C_WHITE)
        cell.border    = _border()
        cell.alignment = _center()
 
    ws.row_dimensions[row_idx].height = 16
    ws.freeze_panes = "C3"
 
    wb.save(out_path)
    print(f"✓ Fichier sauvegardé : {out_path}")



def table_pie_chart(df, var, annee_survenance):
    
    if (annee_survenance-1 not in df['annee_survenance'].unique()) and (annee_survenance in df['annee_survenance'].unique()):
        data = df[df['annee_survenance'] == annee_survenance]
    elif (annee_survenance-1 in df['annee_survenance'].unique()) and (annee_survenance in df['annee_survenance'].unique()):
        data = df[df['annee_survenance'] >= annee_survenance-1]
    else:
        print("La survenance spécifiée n'existe pas dans les données.")
        return None
    table = pd.pivot_table(data, values=var, index='type_beneficiaire',columns='annee_survenance', aggfunc='sum')
    table.index = [f"{el.capitalize()}s" for el in table.index]
    # Calcul des pourcentages
    pt_pct = table.div(table.sum(axis=0), axis=1) * 100

    return pt_pct


def pie_chart(pt_pct, Emplacement_stockage):
    colors = [
        "#173A64",  # bleu
        "#662064",  # vert
        "#D86173",  # jaune
    ]

    n_cols = pt_pct.shape[1]

    if n_cols == 0:
        raise ValueError("La table ne contient aucune colonne")

    if n_cols > 2:
        raise ValueError("La fonction gère uniquement 1 ou 2 années")

    fig, ax = plt.subplots(figsize=(6, 6))

    # === CAS 1 : UNE SEULE ANNÉE ===
    if n_cols == 1:
        col = pt_pct.columns[0]

        ax.pie(
            pt_pct[col],
            radius=1.0,
            labels=None,
            colors=colors,
            autopct='%1.0f%%',
            pctdistance=0.75,
            textprops=dict(color="white", fontweight='bold'),
            wedgeprops=dict(width=0.4, edgecolor='white')
        )

        # Cercle central
        centre_circle = plt.Circle((0, 0), 0.6, fc='white')
        ax.add_artist(centre_circle)

        # Année au centre
        ax.text(
            0, 0,
            col,
            ha='center',
            va='center',
            fontweight='bold'
        )

    # === CAS 2 : DEUX ANNÉES ===
    else:
        # Anneau extérieur (année N)
        ax.pie(
            pt_pct[pt_pct.columns[1]],
            radius=1.0,
            labels=None,
            colors=colors,
            autopct='%1.0f%%',
            pctdistance=0.85,
            textprops=dict(color="white", fontweight='bold'),
            wedgeprops=dict(width=0.3, edgecolor='white')
        )

        # Anneau intérieur (année N-1)
        ax.pie(
            pt_pct[pt_pct.columns[0]],
            radius=0.7,
            labels=None,
            colors=colors,
            autopct='%1.0f%%',
            pctdistance=0.75,
            textprops=dict(color="white", fontweight='bold'),
            wedgeprops=dict(width=0.3, edgecolor='white')
        )

        # Cercle blanc central
        centre_circle = plt.Circle((0, 0), 0.4, fc='white')
        ax.add_artist(centre_circle)

        # Annotation années
        ax.text(0, 1.1, pt_pct.columns[1], ha='center', fontweight='bold',color="#173A64")
        ax.text(0, 0, pt_pct.columns[0], ha='center', fontweight='bold',color="#173A64")

    title="Répartition de la consommation santé par type\nde bénéficiaire au cours des survenances N"

    # Titre
    t = ax.set_title(title)
    t.set_color("#173A64")

    # Légende
    leg = ax.legend(
        pt_pct.index,
        loc='lower center',
        bbox_to_anchor=(0.5, -0.1), ncol=3
    )

    for txt in leg.get_texts():
        txt.set_color("#173A64")
        
    title=title.replace(' ','_').replace('\n','')

    fig.savefig(Emplacement_stockage + '/' + title + str(pt_pct.columns[1]) + '.jpg',
                bbox_inches='tight')

    return fig 